from argparse import ArgumentParser
from decimal import Decimal
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill


class PrecisionWorkbook:

    def __init__(self, output_file, precision, operation):
        self.work_book = Workbook()
        self.work_book.remove_sheet(self.work_book.active)
        self.prec_ws = self.work_book.create_sheet("precision")
        self.lmfao_ws = self.work_book.create_sheet("LMFAO")
        self.comp_ws = self.work_book.create_sheet(operation)
        self.precision = precision
        self.output_file = output_file

    def save_entry(self, row, col, lmfao_val, comp_val):
        diff = abs(lmfao_val - comp_val)
        self.lmfao_ws.cell(row, col).value = lmfao_val
        self.comp_ws.cell(row, col).value = comp_val
        self.prec_ws.cell(row, col).value = diff
        red = openpyxl.styles.colors.Color(rgb='00FF0000')
        red_fill = PatternFill(patternType='solid', fgColor=red)
        green = openpyxl.styles.colors.Color(rgb='00008000')
        green_fill = PatternFill(patternType='solid', fgColor=green)
        self.prec_ws.cell(row, col).fill = red_fill if diff >= self.precision else green_fill

    def save(self):
        self.work_book.save(self.output_file)


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("-lp", "--lmfao_path", dest="lmfao_qr_path", required=True)
    parser.add_argument("-cp", "--comparison_path", dest="comparison_qr_path", required=True)
    parser.add_argument("-op", "--output_file", dest="output_file", required=True)
    parser.add_argument("-pr", "--precision", dest="precision", required=True)
    parser.add_argument("-o", "--operation", dest="operation", required=True)

    args = parser.parse_args()
    lmfao_path = args.lmfao_qr_path
    comparison_path = args.comparison_qr_path
    output_file = args.output_file
    operation = args.operation
    precision = Decimal(args.precision)

    abs_err = Decimal(0)
    abs_err_comp = Decimal(0)
    comp_wb = PrecisionWorkbook(output_file=output_file+".xlsx", precision=precision,
                                operation=operation)
    with open(lmfao_path, 'r') as lmfao_file:
        with open(comparison_path, 'r') as comp_file:
            feats_lmfao = lmfao_file.readline().strip().split(" ")
            feats_comp = comp_file.readline().strip().split(" ")
            row_lmfao = int(feats_lmfao[0])
            col_lmfao = int(feats_lmfao[1]) if (operation == 'qr') else 1
            for row in range(1, row_lmfao + 1):
                lmfao_row_line_val = lmfao_file.readline().strip().split(" ")
                comp_row_line_val = comp_file.readline().strip().split(" ")
                for col in range (0, col_lmfao):
                    lmfao_val_dec = Decimal(lmfao_row_line_val[col])
                    comp_val_dec = Decimal(comp_row_line_val[col])
                    diff = lmfao_val_dec - comp_val_dec
                    abs_err += diff * diff
                    #print(abs_err)
                    abs_err_comp += comp_val_dec * comp_val_dec
                    #print(abs_err_comp)
                    comp_wb.save_entry(row + 1, col + 1, lmfao_val_dec, comp_val_dec)

    comp_wb.save()
    with open(output_file +".txt", 'w') as file_prec:
        file_prec.write("Absolute error is: {}\n".format(abs_err.sqrt()))
        file_prec.write("Frobenius norm of comp is: {}\n".format(abs_err_comp.sqrt()))
        file_prec.write("Relative error is: {}\n".format((abs_err).sqrt() / abs_err_comp.sqrt()))






