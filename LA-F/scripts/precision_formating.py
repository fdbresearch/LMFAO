from argparse import ArgumentParser
from pathlib import Path
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook, load_workbook
import os
import re

reg_pattern = "([\w ]+)(?:[ :]*)([-+]?[0-9]*\.?[0-9]*([eE]?[-+]?[0-9]*))"


def rename_files(comparison_path: str):
    for exp in os.listdir(comparison_path):
        exp_path = os.path.join(comparison_path, exp)
        if not os.path.isdir(exp_path):
            continue
        for file_name in os.listdir(exp_path):
            print(file_name)
            if file_name.endswith(".xlsx") or file_name.endswith(".txt"):
                old_file_name = file_name
                new_file_name = file_name.replace("comp", "comp-")
                new_file_name, sub_made = re.subn("(svd.*)", r"-\1", new_file_name)
                if sub_made == 0:
                    new_file_name = re.sub("(qr.*)", r"-\1", new_file_name)
                print("Old:{} New:{}".format(old_file_name, new_file_name))
                old_file_name = os.path.join(exp_path, old_file_name)
                new_file_name = os.path.join(exp_path, new_file_name)
                os.rename(old_file_name, new_file_name)


class PrecisionWorkbook:
    def __init__(self, prec_file_for, data_op):
        path_prec_file_for = Path(prec_file_for)
        if path_prec_file_for.is_file():
            self.work_book = load_workbook(path_prec_file_for)
            if data_op in self.work_book.sheetnames:
                self.time_sheet = self.work_book[data_op]
            else:
                self.time_sheet = self.work_book.create_sheet(data_op)
        else:
            self.work_book = Workbook()
            self.work_book.remove_sheet(self.work_book.active)
            self.time_sheet = self.work_book.create_sheet(data_op)
        self.output_file = prec_file_for
        self.col_header = 1
        self.row_header = 1

    def save_entry(self, row, col, entry):
        self.time_sheet.cell(row, col).value = entry

    def save_entries(self, data_set, measure_times, data_set_idx):
        self.save_entry(self.row_header, self.col_header, "Measure/dataset")
        self.save_entry(self.row_header, data_set_idx, data_set)

        row_idx = self.row_header + 1
        col_idx = data_set_idx
        #print(data_set_idx)
        for meas_times in measure_times:
            #print(meas_times)
            meas_time = measure_times[meas_times]
            n_times = len(meas_time)
            #print("idx", row_idx)
            for idx, time in enumerate(meas_time, 1):
                self.save_entry(row_idx, col_idx, time)
                self.save_entry(row_idx, self.col_header, meas_times)
                row_idx += 1
            col_letter = get_column_letter(col_idx)
            #print("saved", meas_times, str(row_idx - n_times + 1), str(row_idx - 1))

    def save(self):
        print (self.output_file)
        self.work_book.save(self.output_file)

def parse_precs(precs_path):
    meas_precs = {}
    with open(precs_path) as time_file:
        for line in time_file:
            matches = re.findall(reg_pattern, line)
            if len(matches) > 0:
                meas_name = matches[0][0]
                prec = float(matches[0][1])
                print(meas_name)
                print(prec)

                if meas_name in meas_precs:
                    meas_precs[meas_name].append(prec)
                else:
                    meas_precs[meas_name] = [prec]
    print(meas_precs)
    return meas_precs

if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("-cp", "--comparison_path", dest="comparison_path", required=True)
    parser.add_argument("-r", "--rename", dest="rename", required=True)
    parser.add_argument("-p", "--prec_file_for", dest="prec_file_for", required=True,
                        help='.xlsx file in which prec measurements are stored')

    parser.add_argument("-i", "--precs_dump", dest="precs_path", required=True,
                        help='log file where precs are logged while runtime')
    parser.add_argument("-ds", "--data_set", dest="data_set", required=True)
    parser.add_argument("-op", "--data_op", dest="data_op", required=True)
    parser.add_argument("-s", "--data_set_idx", dest="data_set_idx", required=True)
    args = parser.parse_args()

    comparison_path = args.comparison_path
    prec_file_for = args.prec_file_for
    precs_path = args.precs_path
    rename = True if args.rename == "True" else False
    data_set = args.data_set
    data_op = args.data_op
    data_set_idx = int(args.data_set_idx) + 1

    if rename:
        rename_files(comparison_path)
    measure_precs = parse_precs(precs_path)
    precs_workbook = PrecisionWorkbook(prec_file_for, data_op)
    precs_workbook.save_entries(data_set, measure_precs, data_set_idx)
    precs_workbook.save()











