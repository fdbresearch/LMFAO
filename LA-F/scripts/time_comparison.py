from argparse import ArgumentParser
from pathlib import Path
from openpyxl.utils.cell import get_column_letter
import openpyxl
import re
from openpyxl import Workbook, load_workbook
import os

reg_pattern = "(?<=##LMFAO##)([\w ]+)(?:##)(?:[ ]*)([-+]?[0-9]*\.?[0-9]*([eE]?[-+]?[0-9]*))"

class TimeWorkbook:

    def __init__(self, time_file_for, data_op):
        path_time_file_for = Path(time_file_for)
        if path_time_file_for.is_file():
            self.work_book = load_workbook(path_time_file_for)
            if data_op in self.work_book.sheetnames:
                self.time_sheet = self.work_book[data_op]
            else:
                self.time_sheet = self.work_book.create_sheet(data_op)
        else:
            self.work_book = Workbook()
            self.work_book.remove_sheet(self.work_book.active)
            self.time_sheet = self.work_book.create_sheet(data_op)
        self.output_file = time_file_for
        self.col_header = 1
        self.row_header = 1

    def save_entry(self, row, col, entry):
        self.time_sheet.cell(row, col).value = entry

    def save_entries(self, data_set, measure_times, data_set_idx):
        self.save_entry(self.row_header, self.col_header, "Measure/dataset")
        self.save_entry(self.row_header, data_set_idx, data_set)

        row_idx = self.row_header + 1
        col_idx = data_set_idx
        print(data_set_idx)
        for meas_times in measure_times:
            #print(meas_times)
            meas_time = measure_times[meas_times]
            n_times = len(meas_time)
            self.save_entry(row_idx + len(meas_time), self.col_header,  meas_times)
            print("idx", row_idx)
            for idx, time in enumerate(meas_time, 1):
                self.save_entry(row_idx, col_idx, time)
                self.save_entry(row_idx, self.col_header, meas_times + str(idx))
                row_idx += 1
            col_letter = get_column_letter(col_idx)
            print("saved", meas_times, str(row_idx - n_times + 1), str(row_idx - 1))
            self.save_entry(row_idx, col_idx,
                '=AVERAGE({}{}:{}{})'.format(
                    col_letter,row_idx - n_times + 1,
                    col_letter,row_idx - 1))
            row_idx += 1


    def save(self):
        print (self.output_file)
        self.work_book.save(self.output_file)


def parse_times(times_path, time_workbook):
    meas_times = {}
    with open(times_path) as time_file:
        for line in time_file:
            matches = re.findall(reg_pattern, line)
            if len(matches) > 0:
                #print(line)
                #print(matches)
                measurement = matches[0][0]
                time = float(matches[0][1])
                print(measurement)
                print(time)

                if measurement in meas_times:
                    meas_times[measurement].append(time)
                else:
                    meas_times[measurement] = [time]
    print(meas_times)
    return meas_times

if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("-t", "--time_file_for", dest="time_file_for", required=True,
                        help='.xlsx file in which time measurements are stored')
    parser.add_argument("-i", "--times_path", dest="times_path", required=True,
                        help='log file where times are logged while runtime')
    parser.add_argument("-ds", "--data_set", dest="data_set", required=True)
    parser.add_argument("-op", "--data_op", dest="data_op", required=True)
    parser.add_argument("-s", "--data_set_idx", dest="data_set_idx", required=True)
    args = parser.parse_args()
    data_set = args.data_set
    time_file_for = args.time_file_for
    times_path = args.times_path
    data_op = args.data_op
    data_set_idx = int(args.data_set_idx) + 1

    time_workbook = TimeWorkbook(time_file_for, data_op)

    data_set_times = {}
    measure_times = parse_times(times_path, time_file_for)
    time_workbook.save_entries(data_set, measure_times, data_set_idx)
    time_workbook.save()

